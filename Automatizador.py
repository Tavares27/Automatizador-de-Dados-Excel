import os
from openpyxl import load_workbook

# Verificar se o arquivo existe
caminho = r'C:\Users\Tavares\OneDrive\Documentos\GitHub\Automatizador-de-Dados-Excel\Teste.xlsx'
if not os.path.exists(caminho):
    raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

# Carregar a planilha
arquivo = load_workbook(caminho)
aba_atual = arquivo.active

# Desmesclar todas as células mescladas da aba atual
for merge_range in list(aba_atual.merged_cells.ranges):
    aba_atual.unmerge_cells(str(merge_range))

# Inicializar uma matriz 16x8
matriz = [[0 for _ in range(8)] for _ in range(16)]

# Preencher a matriz com os dados da planilha
for i in range(16):  # Itera pelas linhas
    for j in range(8):  # Itera pelas colunas
        valor = aba_atual.cell(row=(i + 1), column=(j + 1)).value  # Pega o valor da célula
        matriz[i][j] = valor if valor is not None else 0  # Preenche com 0 se a célula estiver vazia

# Inicializar listas GS1 a GS9
gs_listas = [[] for _ in range(9)]  # Cria 9 listas vazias

# Faixas de valor (thresholds)
thresholds = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]

# Processar a matriz e classificar coordenadas em GS1 a GS9
for i in range(1, 16):  # Ignorar a primeira linha (i = 0)
    for j in range(1, 8):  # Ignorar a primeira coluna (j = 0)
        for idx, threshold in enumerate(thresholds):
            if matriz[i][j] <= threshold:
                # Usar os valores da primeira linha e primeira coluna como "etiquetas"
                coordenada = (matriz[i][0], matriz[0][j])  # (valor da primeira coluna, valor da primeira linha)
                gs_listas[idx].append(coordenada)
                break  # Parar após encontrar a faixa correspondente

# Adicionar mensagens de depuração
for idx, gs in enumerate(gs_listas, start=1):
    print(f"GS{idx}: {gs}")

# Escrever os vetores GS na planilha
coluna_inicial = 10  # Coluna onde começar a escrever (ajuste conforme necessário)
for idx, gs in enumerate(gs_listas, start=1):
    # Adicionar um título para cada vetor
    aba_atual.cell(row=1, column=coluna_inicial + idx - 1, value=f"GS{idx}")

    
    # Adicionar os valores de cada GS na planilha
    for linha, coordenada in enumerate(gs, start=2):
        aba_atual.cell(row=linha, column=coluna_inicial + idx - 1, value=str(coordenada))
        
# Salvar as alterações na planilha
arquivo.save(caminho)
print("Vetores GS escritos na planilha com sucesso!")

